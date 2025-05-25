import os
from openpyxl import load_workbook

def modify_excel_file(directory, filename, theme_number, cell_address, yyyymm):
    file_path = os.path.join(directory, filename)
    if not os.path.exists(file_path):
        return False, f"ファイルが見つかりません: {file_path}"

    try:
        wb = load_workbook(file_path)
        ws = wb.active
        ws[cell_address] = theme_number

        new_filename = f"{theme_number}_変更なし_{yyyymm}.xlsx"
        new_filepath = os.path.join(directory, new_filename)
        wb.save(new_filepath)

        return True, new_filepath
    except Exception as e:
        return False, f"エラー: {e}"
